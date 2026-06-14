from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict
from pathlib import Path
from typing import Any
import csv
import json
import html

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    openpyxl = None

from .models import Issue, PageInfo, RunManifest, Hit


ISSUE_FIELDS = list(Issue(
    issue_id="", rule_id="", subject="", message=""
).to_dict().keys())


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_issue_log(out_dir: Path, issues: list[Issue]) -> None:
    write_csv(out_dir / "issue_log.csv", [i.to_dict() for i in issues], ISSUE_FIELDS)


def write_backcheck_log(out_dir: Path, issues: list[Issue]) -> None:
    fields = ISSUE_FIELDS + ["reviewer_decision", "date_closed"]
    rows = []
    for i in issues:
        d = i.to_dict()
        d.setdefault("reviewer_decision", "")
        d.setdefault("date_closed", "")
        rows.append(d)
    write_csv(out_dir / "backcheck_issue_log.csv", rows, fields)


def write_critical_log(out_dir: Path, issues: list[Issue]) -> None:
    rows = [i.to_dict() for i in issues if i.severity in {"Critical", "Major"}]
    write_csv(out_dir / "critical_major_issues.csv", rows, ISSUE_FIELDS)


def write_page_quality(out_dir: Path, page_infos: list[PageInfo]) -> None:
    write_csv(out_dir / "page_quality_report.csv", [asdict(p) for p in page_infos])
    rows = [asdict(p) for p in page_infos if p.raster_only or p.word_count < 20 or p.garbled_text_warning]
    write_csv(out_dir / "ocr_needed_pages.csv", rows)


def write_summary_reports(out_dir: Path, issues: list[Issue], page_infos: list[PageInfo]) -> None:
    sev = Counter(i.severity for i in issues)
    rules = Counter(i.rule_id for i in issues)
    disc = Counter(i.discipline for i in issues)
    sheets = Counter(i.sheet_number for i in issues)

    write_csv(out_dir / "summary_by_severity.csv", [{"severity": k, "count": v} for k, v in sev.most_common()])
    write_csv(out_dir / "summary_by_rule.csv", [{"rule_id": k, "count": v} for k, v in rules.most_common()])
    write_csv(out_dir / "discipline_dashboard.csv", [{"discipline": k, "count": v} for k, v in disc.most_common()])

    weights = {"Critical": 10, "Major": 5, "Minor": 2, "Info": 1}
    by_sheet = defaultdict(list)
    for i in issues:
        by_sheet[i.sheet_number].append(i)
    sheet_rows = []
    for sheet, rows in by_sheet.items():
        score = sum(weights.get(r.severity, 1) for r in rows)
        sheet_rows.append({
            "sheet_number": sheet,
            "risk_score": score,
            "issue_count": len(rows),
            "critical": sum(1 for r in rows if r.severity == "Critical"),
            "major": sum(1 for r in rows if r.severity == "Major"),
            "minor": sum(1 for r in rows if r.severity == "Minor"),
            "info": sum(1 for r in rows if r.severity == "Info"),
        })
    sheet_rows.sort(key=lambda x: x["risk_score"], reverse=True)
    write_csv(out_dir / "sheet_risk_score.csv", sheet_rows)

    combined = []
    for name, counter in [("severity", sev), ("rule", rules), ("discipline", disc), ("sheet", sheets)]:
        for key, count in counter.items():
            combined.append({"group": name, "name": key, "count": count})
    write_csv(out_dir / "summary.csv", combined)


def write_reference_reports(out_dir: Path, issues: list[Issue]) -> None:
    missing: list[Issue] = []
    extra: list[Issue] = []
    dangling: list[Issue] = []
    for i in issues:
        if "listed but not found" in i.subject.lower():
            missing.append(i)
        if "not in reference" in i.subject.lower():
            extra.append(i)
        if "reference target sheet not found" in i.subject.lower():
            dangling.append(i)
    write_csv(out_dir / "reference_only_missing_items.csv", [i.to_dict() for i in missing], ISSUE_FIELDS)
    write_csv(out_dir / "drawing_only_extra_items.csv", [i.to_dict() for i in extra], ISSUE_FIELDS)
    write_csv(out_dir / "dangling_callout_report.csv", [i.to_dict() for i in dangling], ISSUE_FIELDS)
    write_csv(out_dir / "hyperlink_suggestions.csv", [
        {
            "source_sheet": i.sheet_number,
            "source_page": i.page_number,
            "found_reference": i.found_text,
            "target_sheet": "",
            "action": "Verify target sheet then create/link in Bluebeam Batch Link workflow",
            "issue_id": i.issue_id,
        }
        for i in dangling
    ])


def write_powerbi_and_planner(out_dir: Path, issues: list[Issue]) -> None:
    rows = []
    for i in issues:
        rows.append({
            "Issue ID": i.issue_id,
            "Title": i.subject,
            "Description": i.message,
            "Priority": i.severity,
            "Bucket": i.discipline,
            "Assigned To": i.owner,
            "Due Date": i.due_date,
            "Status": i.status,
            "Sheet": i.sheet_number,
            "RFI Candidate": i.rfi_candidate,
        })
    write_csv(out_dir / "planner_import.csv", rows)
    write_csv(out_dir / "powerbi_issues.csv", [i.to_dict() for i in issues], ISSUE_FIELDS)


def write_ai_and_rfi_reports(out_dir: Path, issues: list[Issue]) -> None:
    ai_rows = [{"issue_id": i.issue_id, "sheet_number": i.sheet_number, "subject": i.subject, "draft_comment": i.ai_suggested_comment} for i in issues if i.ai_suggested_comment]
    write_csv(out_dir / "ai_comment_drafts.csv", ai_rows)
    rfi_rows = [i.to_dict() for i in issues if i.rfi_candidate == "Yes"]
    write_csv(out_dir / "rfi_candidates.csv", rfi_rows, ISSUE_FIELDS)


def write_asset_register(out_dir: Path, hits: dict[str, list[Hit]]) -> None:
    rows = []
    for kind in ["Valve", "Instrument", "Equipment", "Line"]:
        seen = set()
        for h in hits.get(kind, []):
            key = (kind, h.normalized, h.sheet_number)
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "asset_tag": h.text,
                "normalized_tag": h.normalized,
                "asset_type": kind,
                "sheet_number": h.sheet_number,
                "page_number": h.page_number,
                "context": h.context,
                "confidence": h.confidence,
                "notes": "Draft asset candidate from drawing text extraction",
            })
    write_csv(out_dir / "asset_register_draft.csv", rows)


def write_symbol_candidates(out_dir: Path, hits: dict[str, list[Hit]]) -> None:
    rows = []
    for kind, hs in hits.items():
        rows.append({"candidate_type": kind, "count": len(hs), "method": "text/tag extraction", "next_step": "Add template/CV detector if required"})
    write_csv(out_dir / "symbol_detection_candidates.csv", rows)


def write_markdown_report(out_dir: Path, issues: list[Issue], manifest: RunManifest) -> None:
    sev = Counter(i.severity for i in issues)
    rules = Counter(i.rule_id for i in issues)
    sheets = Counter(i.sheet_number for i in issues)
    lines = [
        "# Natural Gas Drawing QA Review Report",
        "",
        f"Input: `{manifest.input}`",
        f"Completed: {manifest.completed_at}",
        f"Pages: {manifest.page_count}",
        f"Draft issues: {len(issues)}",
        "",
        "## Severity Counts",
        "",
    ]
    for k, v in sev.most_common():
        lines.append(f"- {k}: {v}")
    lines += ["", "## Top Rules", ""]
    for k, v in rules.most_common(15):
        lines.append(f"- {k}: {v}")
    lines += ["", "## Top Sheets", ""]
    for k, v in sheets.most_common(15):
        lines.append(f"- {k}: {v}")
    lines += ["", "## Critical/Major Issues", ""]
    for i in [x for x in issues if x.severity in {"Critical", "Major"}][:100]:
        lines.append(f"- **{i.issue_id}** `{i.sheet_number}` {i.subject}: {i.message}")
    lines += ["", "## Notes", "", "All findings are draft automated findings and require engineering review."]
    (out_dir / "review_report.md").write_text("\n".join(lines), encoding="utf-8")


def write_html_report(out_dir: Path, issues: list[Issue]) -> None:
    rows = "\n".join(
        f"<tr><td>{html.escape(i.issue_id)}</td><td>{html.escape(i.severity)}</td><td>{html.escape(i.sheet_number)}</td><td>{html.escape(i.rule_id)}</td><td>{html.escape(i.subject)}</td><td>{html.escape(i.message)}</td></tr>"
        for i in issues
    )
    doc = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<title>Natural Gas Drawing QA Dashboard</title>
<style>
body{{font-family:Segoe UI,Arial,sans-serif;margin:24px;}}
table{{border-collapse:collapse;width:100%;font-size:13px;}}
th,td{{border:1px solid #ddd;padding:6px;vertical-align:top;}}
th{{background:#f3f3f3;position:sticky;top:0;}}
.badge{{display:inline-block;padding:2px 6px;border-radius:4px;background:#eee;}}
</style>
</head>
<body>
<h1>Natural Gas Drawing QA Dashboard</h1>
<p><span class="badge">{len(issues)} draft issues</span></p>
<table>
<thead><tr><th>ID</th><th>Severity</th><th>Sheet</th><th>Rule</th><th>Subject</th><th>Message</th></tr></thead>
<tbody>{rows}</tbody>
</table>
</body>
</html>"""
    (out_dir / "review_dashboard.html").write_text(doc, encoding="utf-8")


def write_excel_report(out_dir: Path, issues: list[Issue]) -> None:
    if openpyxl is None:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Issues"
    fields = ISSUE_FIELDS
    ws.append(fields)
    for i in issues:
        d = i.to_dict()
        ws.append([d.get(f, "") for f in fields])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"
    for col_idx, field in enumerate(fields, start=1):
        width = min(60, max(12, len(field) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws2 = wb.create_sheet("Summary")
    sev = Counter(i.severity for i in issues)
    ws2.append(["Severity", "Count"])
    for k, v in sev.most_common():
        ws2.append([k, v])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    wb.save(out_dir / "issue_dashboard.xlsx")


def write_manifest(out_dir: Path, manifest: RunManifest) -> None:
    (out_dir / "run_manifest.json").write_text(json.dumps(manifest.__dict__, indent=2), encoding="utf-8")


def write_rules_documentation(out_dir: Path, config: dict[str, Any]) -> None:
    lines = ["# Active QA Rules", ""]
    for rule_id, cfg in sorted(config.get("rules", {}).items()):
        lines.append(f"## {rule_id}")
        lines.append(f"- Enabled: {cfg.get('enabled', True)}")
        lines.append(f"- Severity: {cfg.get('severity', 'Info')}")
        lines.append(f"- Discipline: {cfg.get('discipline', 'General')}")
        lines.append(f"- Confidence: {cfg.get('confidence', '')}")
        lines.append("")
    (out_dir / "rules_documentation.md").write_text("\n".join(lines), encoding="utf-8")


def write_integration_manifests(out_dir: Path, config: dict[str, Any]) -> None:
    docs = {
        "bluebeam_bax_notes.md": """# Bluebeam BAX Integration Notes

This release uses standard PDF annotations because they are stable and visible in Bluebeam.

Future BAX implementation path:
1. Create canonical Bluebeam markups manually.
2. Export them as `.bax`.
3. Inspect template structure.
4. Build a controlled BAX template writer.
5. Validate import behavior per Bluebeam version before production use.
""",
        "studio_api_notes.md": """# Bluebeam Studio/API Integration Notes

This release outputs CSV/Excel/HTML artifacts suitable for review governance.

Future Studio/API path:
1. Register Bluebeam developer application.
2. Authenticate through the Studio API.
3. Create session/project workflow.
4. Upload marked-up PDF.
5. Sync markup status changes to dashboard/ticket system.
""",
        "field_photo_matching_manifest.md": """# Field Photo-to-Drawing Matching Hook

This release does not perform photo matching.

Recommended future workflow:
1. Capture field photos with filenames tied to station/asset/tag.
2. Extract EXIF timestamp/GPS when allowed.
3. Match against asset_register_draft.csv and drawing sheet references.
4. Produce photo-to-drawing review pack.
""",
    }
    for name, content in docs.items():
        (out_dir / name).write_text(content, encoding="utf-8")


def write_all_reports(
    out_dir: Path,
    issues: list[Issue],
    page_infos: list[PageInfo],
    hits: dict[str, list[Hit]],
    manifest: RunManifest,
    config: dict[str, Any],
) -> None:
    write_issue_log(out_dir, issues)
    write_backcheck_log(out_dir, issues)
    write_critical_log(out_dir, issues)
    write_page_quality(out_dir, page_infos)
    write_summary_reports(out_dir, issues, page_infos)
    write_reference_reports(out_dir, issues)
    write_powerbi_and_planner(out_dir, issues)
    write_ai_and_rfi_reports(out_dir, issues)
    write_asset_register(out_dir, hits)
    write_symbol_candidates(out_dir, hits)
    write_markdown_report(out_dir, issues, manifest)
    write_html_report(out_dir, issues)
    write_excel_report(out_dir, issues)
    write_rules_documentation(out_dir, config)
    write_integration_manifests(out_dir, config)
    write_manifest(out_dir, manifest)
