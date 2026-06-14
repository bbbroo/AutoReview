from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import csv
import re

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


@dataclass
class ReferenceRecord:
    normalized: str
    raw: str
    source_type: str
    source_file: str
    row_number: int
    fields: dict[str, str]


def normalize_value(value: str, fuzzy: bool = True, remove_separators: bool = False) -> str:
    value = str(value or "").strip().upper()
    value = value.replace("_", "-")
    value = re.sub(r"\s+", "-", value)
    value = re.sub(r"-+", "-", value)
    if fuzzy and remove_separators:
        return re.sub(r"[-\s_\"']", "", value)
    return value


def pick_column(headers: list[str], candidates: list[str]) -> str | None:
    normalized = {h.strip().lower().replace(" ", "_"): h for h in headers if h}
    for candidate in candidates:
        key = candidate.strip().lower().replace(" ", "_")
        if key in normalized:
            return normalized[key]
    return None


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def _read_xlsx(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for .xlsx support. Install with: python -m pip install openpyxl")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c or "").strip() for c in rows[0]]
    out = []
    for row in rows[1:]:
        out.append({headers[i]: str(row[i] if i < len(row) and row[i] is not None else "").strip() for i in range(len(headers))})
    return headers, out


def read_table(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError(f"Unsupported reference file type: {path}. Use .csv or .xlsx")


def load_reference_records(
    path: str | Path | None,
    source_type: str,
    column_map: dict[str, list[str]],
    config: dict[str, Any],
) -> list[ReferenceRecord]:
    if not path:
        return []
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Reference file not found: {path}")

    headers, rows = read_table(path)
    tag_col = pick_column(headers, column_map.get("tag", []) or column_map.get("sheet_number", []))
    if not tag_col:
        raise ValueError(f"Could not find tag/sheet column in {path}. Headers: {headers}")

    mapped_cols: dict[str, str] = {}
    for field, candidates in column_map.items():
        col = pick_column(headers, candidates)
        if col:
            mapped_cols[field] = col

    norm_cfg = config.get("normalization", {})
    fuzzy = bool(norm_cfg.get("fuzzy_tags", True))
    remove_sep = bool(norm_cfg.get("remove_separators_for_match", False))

    records = []
    for idx, row in enumerate(rows, start=2):
        raw = str(row.get(tag_col, "") or "").strip()
        if not raw:
            continue
        fields = {field: str(row.get(col, "") or "").strip() for field, col in mapped_cols.items()}
        records.append(ReferenceRecord(
            normalized=normalize_value(raw, fuzzy=fuzzy, remove_separators=remove_sep),
            raw=raw,
            source_type=source_type,
            source_file=str(path),
            row_number=idx,
            fields=fields,
        ))
    return records


def load_aliases(path: str | Path | None, config: dict[str, Any]) -> dict[str, str]:
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    headers, rows = read_table(path)
    old_col = pick_column(headers, ["old", "alias", "from", "source", "drawing_tag"])
    new_col = pick_column(headers, ["new", "canonical", "to", "target", "reference_tag"])
    if not old_col or not new_col:
        return {}
    norm_cfg = config.get("normalization", {})
    aliases = {}
    for row in rows:
        old = normalize_value(row.get(old_col, ""), norm_cfg.get("fuzzy_tags", True), norm_cfg.get("remove_separators_for_match", False))
        new = normalize_value(row.get(new_col, ""), norm_cfg.get("fuzzy_tags", True), norm_cfg.get("remove_separators_for_match", False))
        if old and new:
            aliases[old] = new
    return aliases


def load_ignore_patterns(path: str | Path | None) -> list[re.Pattern]:
    if not path:
        return []
    path = Path(path)
    if not path.exists():
        return []
    patterns = []
    if path.suffix.lower() == ".csv":
        headers, rows = read_table(path)
        col = pick_column(headers, ["pattern", "regex", "ignore"])
        if col:
            for row in rows:
                p = str(row.get(col, "") or "").strip()
                if p:
                    patterns.append(re.compile(p, re.IGNORECASE))
    else:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                patterns.append(re.compile(line, re.IGNORECASE))
    return patterns


def apply_alias(value: str, aliases: dict[str, str]) -> str:
    return aliases.get(value, value)


def validate_reference_records(records: list[ReferenceRecord]) -> list[str]:
    warnings = []
    seen: dict[tuple[str, str], ReferenceRecord] = {}
    for rec in records:
        key = (rec.source_type, rec.normalized)
        if key in seen:
            warnings.append(f"Duplicate {rec.source_type} reference '{rec.raw}' in {rec.source_file}, rows {seen[key].row_number} and {rec.row_number}.")
        seen[key] = rec
    return warnings
